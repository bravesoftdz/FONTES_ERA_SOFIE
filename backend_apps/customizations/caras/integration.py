# coding: utf-8

"""

"""

from openpyxl import Workbook

from library.storage.config import ConfigStorage
from micro_contract.transaction import Transaction


class BuildExcel(object):
    """

    """

    EMPRESA_CODE = 134

    CODIGO_MY_SOFIE = 2751

    HEADERS_2_FTP = [
        'Empresa_id',
        'Oferta_id',
        'Produto_id',
        'Nome_Completo_txt',
        'Nome_Fantasia_txt',
        'Aos_Cuidados_Txt',
        'Email_txt',
        'Email_Com_txt',
        'CpfCgc_txt',
        'RG_InscricaoEstadual_txt',
        'Cod_Cli_Terceiro',
        'Cod_Ass_Terceiro',
        'DDI_num',
        'DDD_num',
        'Fone_txt',
        'DDI_1_num',
        'DDD_1_num',
        'Fone_1_txt',
        'DDI_Cel_num',
        'DDD_Cel_num',
        'Fone_Cel_txt',
        'DDI_Fax_num',
        'DDD_Fax_num',
        'Fone_Fax_txt',
        'Representante_id',
        'Sexo_txt',
        'Pessoa_txt',
        'CEP_num',
        'Endereco_txt',
        'Numero_txt',
        'Complemento_txt',
        'Referencia_txt',
        'Bairro_txt',
        'Cidade_txt',
        'Estado_txt',
        'Pais_id',
        'Pais_txt',
        'ZIPCode_txt',
        'flg_Exterior',
        'Dia_Nasc_num',
        'Mês_Nasc_Num',
        'Ano_Nasc_Num',
        'Site_txt',
        'Formato_txt',
        'Origem_txt',
        'Banner_txt',
        'Versao_txt',
        'Nome_Vendedor_txt',
        'Ponto_Venda_txt',
        'Quantidade_Exemplares',
    ]

    def build_row(self, transaction: Transaction) -> dict:
        """

        :return:
        """
        ano, mes, dia, *_ = transaction.consumer.birthday.timetuple()

        return {
            'Empresa_id': BuildExcel.EMPRESA_CODE,
            'Oferta_id': transaction.card.reference_codes[2],
            'Produto_id': transaction.card.reference_codes[1],
            'Nome_Completo_txt': transaction.consumer.full_name,
            'Nome_Fantasia_txt': '',
            'Aos_Cuidados_Txt': transaction.consumer.to_care,
            'Email_txt': transaction.consumer.email,
            'Email_Com_txt': '',
            'CpfCgc_txt': transaction.consumer.consumer,
            'RG_InscricaoEstadual_txt': '',
            'Cod_Cli_Terceiro': '',
            'Cod_Ass_Terceiro': '',
            'DDI_num': '55',
            'DDD_num': transaction.consumer.phone[0:2],
            'Fone_txt': transaction.consumer.phone[2:],
            'DDI_1_num': '',
            'DDD_1_num': '',
            'Fone_1_txt': '',
            'DDI_Cel_num': '',
            'DDD_Cel_num': '',
            'Fone_Cel_txt': '',
            'DDI_Fax_num': '',
            'DDD_Fax_num': '',
            'Fone_Fax_txt': '',
            'Representante_id': BuildExcel.CODIGO_MY_SOFIE,
            'Sexo_txt': transaction.consumer.gender,
            'Pessoa_txt': transaction.consumer.kind_of_person,
            'CEP_num': transaction.address.zipcode,
            'Endereco_txt': '{} {}'.format(transaction.address.type, transaction.address.full_name),
            'Numero_txt': transaction.address.number,
            'Complemento_txt': transaction.address.complement,
            'Referencia_txt': transaction.address.reference,
            'Bairro_txt': transaction.address.district,
            'Cidade_txt': transaction.address.city,
            'Estado_txt': transaction.address.state,
            'Pais_id': '',
            'Pais_txt': 'Brasil',
            'ZIPCode_txt': '',
            'flg_Exterior': '',
            'Dia_Nasc_num': dia,
            'Mês_Nasc_Num': mes,
            'Ano_Nasc_Num': ano,
            'Site_txt': '',
            'Formato_txt': '',
            'Origem_txt': '',
            'Banner_txt': '',
            'Versao_txt': '',
            'Nome_Vendedor_txt': transaction.sofier.full_name.upper(),
            'Ponto_Venda_txt': '',
            'Quantidade_Exemplares': 1
        }

    def __call__(self):
        """
        Referências:
            - https://stackoverflow.com/questions/12976378/openpyxl-convert-csv-to-excel

        :return:
        """
        cursor = ConfigStorage().config['transaction'].find({'actors.company': 'caras', 'status.status': 'FINISHED', 'status.success': True}, {'transaction': 1})

        all_rows = list()
        for item in cursor:
            trans = Transaction(item['transaction'])
            one_row = self.build_row(trans)
            all_rows.append(one_row)

        if len(all_rows) > 0:
            xls_file = 'c:\\teste\\caras.xlsx'

            wb = Workbook()
            ws = wb.worksheets[0]
            ws.title = 'BaseInclusao'

            for col, value in enumerate(BuildExcel.HEADERS_2_FTP, 1):
                ws.cell(1, col, value)

            for row, item in enumerate(all_rows, 2):
                for col, head in enumerate(BuildExcel.HEADERS_2_FTP, 1):
                    ws.cell(row, col, item.get(head, ''))

            wb.save(filename=xls_file)


if __name__ == '__main__':
    excel = BuildExcel()
    excel()
